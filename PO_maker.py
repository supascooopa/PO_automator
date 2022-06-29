from openpyxl import load_workbook
from openpyxl import Workbook
from file_manager_v101 import get_file_name

excel_file = get_file_name()
read_wb = load_workbook(excel_file)
read_ws = read_wb.active

write_wb = Workbook()
write_ws = write_wb.active

write_ws.append(["phone description", "quantity", "unit price USD"])
for rows in read_ws.iter_rows(min_row=2, values_only=True):
    if "-" not in rows and rows[3] is not None:
        if isinstance(rows[3], str):

            split_strings = rows[3].split(",")
            for c_plus_q in split_strings:
                quantity = c_plus_q.split()[0]
                color = c_plus_q.split()[1].strip()
                phone_name = rows[0] + " " + rows[1] + " " + color
                phone_price = rows[2]
                row_list = [phone_name, quantity, phone_price]
                write_ws.append(row_list)
        elif isinstance(rows[3], int):
            quantity = rows[3]
            try:
                phone_name = rows[0] + rows[1]
            except TypeError:
                phone_name = rows[0]
            phone_price = rows[2]
            row_list = [phone_name, quantity, phone_price]
            write_ws.append(row_list)

write_wb.save("new_po.xlsx")

