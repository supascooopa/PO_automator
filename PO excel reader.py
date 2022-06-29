import openpyxl
import file_manager_v101

file = file_manager_v101.get_file_name(file_extension=".xlsx")
wb = openpyxl.load_workbook(file)
sheets_lst = []
for index, sheets in enumerate(wb.sheetnames):
    print(sheets)
    sheets_lst.append(sheets)
sheets_question = int(input("Please write the corresponding number of the sheet you want to work on: "))
ws = wb[sheets_lst[sheets_question - 1]]

new_wb = openpyxl.Workbook()
new_ws = new_wb.active

# row_for_new_ws = 3
# iterating through the rows
for row in ws.iter_rows(min_row=1, min_col=1):
    # getting model and name of the phone
    model = row[0].value
    description = row[1].value
    price = row[2].value
    print(model, description)
    usr_input = input("please write colors, or don't write any if you don't want to add this phone: ")
    tup_input = tuple(x.strip() for x in usr_input.split(','))
    if tup_input[0]:
        for color in tup_input:
            try:
                model_plus_description_color = model + " " + description + " " + color
                # new_ws.cell(row_for_new_ws, 1).value = model + " " + description + " " + color
                quantity = int(input("enter quantity: "))
                new_ws.append([model_plus_description_color, quantity, price])
                # new_ws.cell(row_for_new_ws, 2).value = quantity
                # row_for_new_ws += 1
            except TypeError:
                model_color = model + " " + color
                quantity = int(input("enter quantity: "))
                new_ws.append([model_color, quantity, price])

new_file_name = input("Please write a file name: ")
new_wb.save(f"{new_file_name}.xlsx")




